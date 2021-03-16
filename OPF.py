import PySimpleGUI as sg 
import openpyxl 

sg.theme('DarkGrey6')

#=========================================================>
#Выбор документа в форме PySimple
#def open_file():
#    sg.set_options(auto_size_buttons=True)
#    openfile = sg.popup_get_file(
#         "Открыть документ",
#          title="ОПФ",
#          file_types=(("XLSX Files", "*.xlsx"),("CSV Files", "*.csv")))
#    
#    if openfile == '':
#        return
#=========================================================>


#book = openpyxl.open(openfile, read_only=True)             # подключаем выбранный файл
book = openpyxl.open('Оценка ОПФ.xlsx', read_only=True)

sheet_first = book.worksheets[0]                            #Работаем с первым листом
sheet_second = book.worksheets[1]                           #Работаем со вторым листом
sheet_third = book.worksheets[2]                            #Работаем с третьим листом
sheet_fourth = book.worksheets[3]                           #Работаем с четвертым листом
sheet_fifth = book.worksheets[4]                            #Работаем с пятым листом

header_list_first = []
data_list_first = []

for column in range(sheet_first.min_column - 1, sheet_first.max_column):  # Расчехляем значения заголовков
    if sheet_first[1][column].value is not None:
        header_list_first.append(sheet_first[1][column].value)
print(header_list_first)

for row in sheet_first.iter_rows(sheet_first.min_row + 1, sheet_first.max_row):  # Расчехляем значения данных
    for cell in row:
        if cell.value is not None:
            data_list_first.append(cell.value)


#===============>
data = [['rocsptjach', 161, 570, 844, 745, 454],
['jwsqgvyatn', 380, 524, 697, 124, 911],
['egeflqdyvd', 813, 138, 834, 292, 625]]


layout = [[sg.Table(values=data_list_first, headings=header_list_first, max_col_width=35,
                    # background_color='light blue',
                    auto_size_columns=True,
                    display_row_numbers=True,
                    justification='right',
                    num_rows=20,
                    alternating_row_color='lightyellow',
                    key='-TABLE-',
                    row_height=35,
                    tooltip='This is a table')],
          [sg.Button('Read')],
          [sg.Text('Read = read which rows are selected')]]

# ------ Create Window ------
window = sg.Window('The Table Element', layout)

# ------ Event Loop ------
while True:
    event, values = window.read()
    print(event, values)
    if event == sg.WIN_CLOSED:
        break

window.close()

#===============>


# header_list_first = [sheet_first[1][0].value, sheet_first[1][1].value,
#                      sheet_first[1][2].value, sheet_first[1][3].value,
#                      sheet_first[1][4].value, sheet_first[1][5].value]
# print(header_list_first)








    



#if __name__ == '__main__':
#    open_file()

                            